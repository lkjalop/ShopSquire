"""Bounded, tenant-scoped fetching of credential-free official market sources.

Only registry-approved HTTPS origins and parsers are callable. Raw provider
text is never treated as SKU exposure or execution authority.
"""
from __future__ import annotations

import hashlib
import io
import json
import os
import re
import uuid
import csv
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any, Callable
from urllib.parse import urlparse

from sqlalchemy import text

from src.app.services.market_evidence_policy import resolve_contradictions
from src.app.services.market_source_registry import (
    govern_external_observation,
    load_market_source_registry,
)


_TABLE = "market_source_fetch_revision"
_MAX_ROWS = 100


@dataclass(frozen=True)
class PublicHttpResponse:
    status_code: int
    headers: dict[str, str]
    body: bytes


Transport = Callable[
    [str, dict[str, str], dict[str, str], float],
    PublicHttpResponse,
]


def _utc(value: datetime | None = None) -> datetime:
    selected = value or datetime.now(timezone.utc)
    if selected.tzinfo is None:
        selected = selected.replace(tzinfo=timezone.utc)
    return selected.astimezone(timezone.utc)


def _request_key(query: dict[str, str]) -> str:
    return hashlib.sha256(
        json.dumps(query, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _default_transport(
    url: str,
    params: dict[str, str],
    headers: dict[str, str],
    timeout: float,
) -> PublicHttpResponse:
    import httpx

    with httpx.Client(
        timeout=httpx.Timeout(timeout),
        # Registry validation covers the configured origin. Redirects are not
        # followed because they could escape that allowlisted host.
        follow_redirects=False,
        headers={"User-Agent": "ShopSquire-market-intelligence/1.0", **headers},
    ) as client:
        response = client.get(url, params=params) if params else client.get(url)
    return PublicHttpResponse(
        status_code=int(response.status_code),
        headers={str(key).lower(): str(value) for key, value in response.headers.items()},
        body=bytes(response.content),
    )


def _normalize_cpsc_query(query: dict[str, Any]) -> dict[str, str]:
    mapping = {
        "recall_date_start": "RecallDateStart",
        "recall_date_end": "RecallDateEnd",
        "product_name": "ProductName",
        "title": "RecallTitle",
    }
    normalized = {"format": "json"}
    for source_key, provider_key in mapping.items():
        value = str(query.get(source_key) or "").strip()
        if value:
            if len(value) > 120:
                raise ValueError("public_market_query_value_too_long")
            normalized[provider_key] = value
    if len(normalized) == 1:
        raise ValueError("public_market_bounded_filter_required")
    return normalized


def _normalize_world_bank_query(query: dict[str, Any]) -> dict[str, str]:
    raw_series = query.get("series") or []
    if isinstance(raw_series, str):
        raw_series = [part.strip() for part in raw_series.split("|")]
    if not isinstance(raw_series, list):
        raise ValueError("public_market_series_invalid")
    series = sorted({
        str(value or "").strip()
        for value in raw_series
        if str(value or "").strip()
    })
    if not series or len(series) > 5 or any(len(value) > 120 for value in series):
        raise ValueError("public_market_series_invalid")
    signal_type = str(query.get("signal_type") or "commodity_input_price").strip()
    if signal_type not in {
        "commodity_input_price",
        "transport_fuel_price",
        "energy_input_price",
    }:
        raise ValueError("external_market_signal_not_permitted")
    return {"series": "|".join(series), "signal_type": signal_type}


def _normalize_nws_query(query: dict[str, Any]) -> dict[str, str]:
    """Require a bounded geographic scope for the active-alert endpoint."""
    point = str(query.get("point") or "").strip()
    area = str(query.get("area") or "").strip().upper()
    zone = str(query.get("zone") or "").strip().upper()
    selected = sum(bool(value) for value in (point, area, zone))
    if selected != 1:
        raise ValueError("public_market_nws_single_scope_required")
    if point:
        match = re.fullmatch(r"(-?\d{1,2}(?:\.\d{1,6})?),(-?\d{1,3}(?:\.\d{1,6})?)", point)
        if not match:
            raise ValueError("public_market_nws_point_invalid")
        latitude, longitude = (float(match.group(1)), float(match.group(2)))
        if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
            raise ValueError("public_market_nws_point_invalid")
        return {"point": f"{latitude:.6f},{longitude:.6f}"}
    if area:
        if not re.fullmatch(r"[A-Z]{2}", area):
            raise ValueError("public_market_nws_area_invalid")
        return {"area": area}
    if not re.fullmatch(r"[A-Z]{2}[CZ]\d{3}|[A-Z]{2}\d{3}", zone):
        raise ValueError("public_market_nws_zone_invalid")
    return {"zone": zone}


def _bounded_values(
    query: dict[str, Any],
    key: str,
    *,
    maximum: int,
    required: bool = False,
) -> list[str]:
    raw_values = query.get(key) or []
    if isinstance(raw_values, str):
        raw_values = [part.strip() for part in raw_values.split("|")]
    if not isinstance(raw_values, list):
        raise ValueError(f"public_market_{key}_invalid")
    values = sorted({
        str(value or "").strip()
        for value in raw_values
        if str(value or "").strip()
    })
    if (
        (required and not values)
        or len(values) > maximum
        or any(len(value) > 120 for value in values)
    ):
        raise ValueError(f"public_market_{key}_invalid")
    return values


def _normalize_usgs_query(query: dict[str, Any]) -> dict[str, str]:
    commodities = _bounded_values(query, "commodities", maximum=5, required=True)
    statistics = _bounded_values(query, "statistics", maximum=5)
    countries = _bounded_values(query, "countries", maximum=10)
    latest = query.get("latest_year_only", True)
    if not isinstance(latest, bool):
        raise ValueError("public_market_latest_year_only_invalid")
    return {
        "commodities": "|".join(commodities),
        "statistics": "|".join(statistics),
        "countries": "|".join(countries),
        "latest_year_only": "true" if latest else "false",
    }


def _usgs_data_response(
    profile: dict[str, Any],
    *,
    transport: Transport,
    conditional_headers: dict[str, str],
) -> tuple[PublicHttpResponse, dict[str, Any]]:
    """Resolve one exact file from one pinned ScienceBase item.

    The item endpoint is intentionally fetched without conditional data-file
    headers. The discovered file must remain on the pinned item path and host.
    """
    timeout = float(profile.get("timeout_seconds") or 8)
    metadata_response = transport(
        str(profile["url"]),
        {"format": "json"},
        {},
        timeout,
    )
    if metadata_response.status_code < 200 or metadata_response.status_code >= 300:
        return metadata_response, {}
    if len(metadata_response.body) > int(
        profile.get("max_metadata_bytes") or 500_000
    ):
        raise ValueError("public_market_usgs_metadata_too_large")
    try:
        metadata = json.loads(metadata_response.body.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("public_market_usgs_metadata_invalid") from exc
    item_id = str(profile["item_id"])
    file_name = str(profile["csv_name"])
    if not isinstance(metadata, dict) or str(metadata.get("id") or "") != item_id:
        raise ValueError("public_market_usgs_item_identity_invalid")
    files = metadata.get("files")
    matches = [
        item for item in files
        if (
            isinstance(item, dict)
            and str(item.get("name") or "") == file_name
            and str(item.get("contentType") or "").casefold() == "text/csv"
        )
    ] if isinstance(files, list) else []
    if len(matches) != 1:
        raise ValueError("public_market_usgs_file_identity_invalid")
    selected = matches[0]
    try:
        declared_size = int(selected.get("size"))
    except (TypeError, ValueError) as exc:
        raise ValueError("public_market_usgs_file_size_invalid") from exc
    if declared_size < 1 or declared_size > int(
        profile.get("max_response_bytes") or 4_000_000
    ):
        raise ValueError("public_market_usgs_file_size_invalid")
    file_url = str(selected.get("url") or "")
    parsed = urlparse(file_url)
    if (
        parsed.scheme != "https"
        or parsed.hostname != str(profile["allowed_host"]).casefold()
        or parsed.username is not None
        or parsed.password is not None
        or parsed.port is not None
        or parsed.path != f"/catalog/file/get/{item_id}"
    ):
        raise ValueError("public_market_usgs_file_origin_invalid")
    response = transport(file_url, {}, conditional_headers, timeout)
    provenance = metadata.get("provenance")
    return response, {
        "item_id": item_id,
        "release_doi": str(profile["release_doi"]),
        "published_at": (
            str(provenance.get("lastUpdated") or provenance.get("dateCreated") or "")
            if isinstance(provenance, dict)
            else ""
        ),
        "file_name": file_name,
        "declared_size": declared_size,
        "provider_checksum": selected.get("checksum"),
    }


def _usgs_number(raw: str) -> tuple[float | None, str]:
    cleaned = raw.strip()
    estimated = cleaned.casefold().startswith("e")
    candidate = cleaned[1:].strip() if estimated else cleaned
    candidate = candidate.replace(",", "")
    if not re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)", candidate):
        return None, "reported_non_numeric"
    return float(candidate), "estimated_numeric" if estimated else "observed_numeric"


def _usgs_observations(
    body: bytes,
    *,
    request: dict[str, str],
    retrieved_at: str,
    release: dict[str, Any],
) -> list[dict[str, Any]]:
    try:
        decoded = body.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            decoded = body.decode("cp1252")
        except UnicodeDecodeError as exc:
            raise ValueError("public_market_usgs_csv_invalid") from exc
    reader = csv.DictReader(io.StringIO(decoded))
    required = {
        "Commodity",
        "Country",
        "Statistics",
        "Statistics_detail",
        "Unit",
        "Year",
        "Value",
        "Notes",
        "Is critical mineral 2025",
    }
    if not reader.fieldnames or not required <= set(reader.fieldnames):
        raise ValueError("public_market_usgs_csv_shape_invalid")
    commodities = {
        value.casefold() for value in request["commodities"].split("|") if value
    }
    statistics = {
        value.casefold() for value in request["statistics"].split("|") if value
    }
    countries = {
        value.casefold() for value in request["countries"].split("|") if value
    }
    matches: list[dict[str, str]] = []
    for raw_row in reader:
        row = {str(key): str(value or "").strip() for key, value in raw_row.items()}
        if row["Commodity"].casefold() not in commodities:
            continue
        if statistics and row["Statistics"].casefold() not in statistics:
            continue
        if countries and row["Country"].casefold() not in countries:
            continue
        if not re.fullmatch(r"\d{4}", row["Year"]):
            continue
        matches.append(row)
        if len(matches) > 25_000:
            raise ValueError("public_market_usgs_filter_too_broad")
    if request["latest_year_only"] == "true" and matches:
        latest_year = max(int(row["Year"]) for row in matches)
        matches = [row for row in matches if int(row["Year"]) == latest_year]
    published_at = str(release.get("published_at") or retrieved_at)
    observations: list[dict[str, Any]] = []
    for row in matches[:_MAX_ROWS]:
        numeric_value, value_status = _usgs_number(row["Value"])
        identity = "|".join((
            row["Commodity"],
            row["Country"],
            row["Statistics"],
            row["Statistics_detail"],
            row["Year"],
        ))
        record_hash = hashlib.sha256(identity.encode("utf-8")).hexdigest()[:24]
        measurement = {
            "kind": "mineral_commodity_statistic",
            "direction": "unknown",
            "commodity": row["Commodity"],
            "country": row["Country"],
            "statistic": row["Statistics"],
            "statistic_detail": row["Statistics_detail"],
            "year": int(row["Year"]),
            "reported_value": row["Value"],
            "numeric_value": numeric_value,
            "value_status": value_status,
            "uom": row["Unit"] or None,
            "critical_mineral_2025": (
                row["Is critical mineral 2025"].casefold() in {"yes", "true", "1"}
            ),
            "notes": row["Notes"][:1000],
            "release_doi": release.get("release_doi"),
            "sciencebase_item_id": release.get("item_id"),
            "provider_checksum": release.get("provider_checksum"),
        }
        observations.append(govern_external_observation(
            source_id="usgs_minerals",
            source_record_id=f"mcs2026:{record_hash}",
            signal_type="mineral_supply_statistic",
            subject_id=f"usgs-mineral:{row['Commodity'].casefold()}",
            measurement=measurement,
            geography=row["Country"] or "unspecified",
            effective_from=f"{row['Year']}-01-01T00:00:00+00:00",
            effective_to=None,
            published_at=published_at,
            available_at=published_at,
            retrieved_at=retrieved_at,
        ))
    return observations


def _cpsc_observations(
    rows: Any,
    *,
    retrieved_at: str,
) -> list[dict[str, Any]]:
    if not isinstance(rows, list):
        raise ValueError("public_market_response_shape_invalid")
    observations: list[dict[str, Any]] = []
    for row in rows[:_MAX_ROWS]:
        if not isinstance(row, dict):
            continue
        record_id = str(row.get("RecallNumber") or row.get("RecallID") or "").strip()
        recall_date = str(row.get("RecallDate") or "").strip()
        published_at = str(row.get("LastPublishDate") or recall_date).strip()
        if not record_id or not recall_date:
            continue
        products = row.get("Products") if isinstance(row.get("Products"), list) else []
        product_names = sorted({
            str(product.get("Name") or "").strip()
            for product in products
            if isinstance(product, dict) and str(product.get("Name") or "").strip()
        })
        subject = product_names[0] if product_names else str(row.get("Title") or record_id)
        measurement = {
            "kind": "product_recall",
            "direction": "adverse",
            "recall_number": record_id,
            "title": str(row.get("Title") or "")[:500],
            "products": product_names[:20],
            "hazards": [
                str(hazard.get("Name") or "").strip()
                for hazard in (row.get("Hazards") or [])
                if isinstance(hazard, dict) and str(hazard.get("Name") or "").strip()
            ][:20],
            "source_url": str(row.get("URL") or "")[:1000],
        }
        observations.append(govern_external_observation(
            source_id="cpsc_recalls",
            source_record_id=record_id,
            signal_type="product_recall",
            subject_id=f"cpsc-product:{subject.casefold()}",
            measurement=measurement,
            geography="US",
            effective_from=recall_date,
            effective_to=None,
            published_at=published_at,
            available_at=published_at,
            retrieved_at=retrieved_at,
        ))
    return observations


def _nws_observations(
    payload: Any,
    *,
    request: dict[str, str],
    retrieved_at: str,
) -> list[dict[str, Any]]:
    if not isinstance(payload, dict) or not isinstance(payload.get("features"), list):
        raise ValueError("public_market_response_shape_invalid")
    observations: list[dict[str, Any]] = []
    for feature in payload["features"][:_MAX_ROWS]:
        if not isinstance(feature, dict):
            continue
        properties = feature.get("properties")
        if not isinstance(properties, dict):
            continue
        record_id = str(feature.get("id") or properties.get("id") or "").strip()
        published = str(properties.get("sent") or "").strip()
        effective = str(
            properties.get("effective") or properties.get("onset") or published
        ).strip()
        if not record_id or not published or not effective:
            continue
        geocode = properties.get("geocode") if isinstance(properties.get("geocode"), dict) else {}
        ugc = geocode.get("UGC") if isinstance(geocode.get("UGC"), list) else []
        scope_ref = str(ugc[0]).strip() if ugc else next(iter(request.values()))
        geography = str(properties.get("areaDesc") or scope_ref).strip()
        measurement = {
            "kind": "weather_alert",
            "direction": "adverse",
            "event": str(properties.get("event") or "")[:200],
            "severity": str(properties.get("severity") or "unknown")[:40],
            "certainty": str(properties.get("certainty") or "unknown")[:40],
            "urgency": str(properties.get("urgency") or "unknown")[:40],
            "message_type": str(properties.get("messageType") or "unknown")[:40],
            "status": str(properties.get("status") or "unknown")[:40],
            "headline": str(properties.get("headline") or "")[:500],
            "source_url": record_id[:1000],
            "supply_chain_stage": "transport_lane",
        }
        observations.append(govern_external_observation(
            source_id="nws_weather_alerts",
            source_record_id=record_id,
            signal_type="lane_weather_risk",
            subject_id=f"nws-zone:{scope_ref.casefold()}",
            measurement=measurement,
            geography=geography,
            effective_from=effective,
            effective_to=(
                str(properties.get("ends") or properties.get("expires") or "").strip()
                or None
            ),
            published_at=published,
            available_at=published,
            retrieved_at=retrieved_at,
        ))
    return observations


def _world_bank_observations(
    body: bytes,
    *,
    request: dict[str, str],
    retrieved_at: str,
) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
        sheet = workbook["Monthly Prices"]
    except Exception as exc:
        raise ValueError("public_market_workbook_invalid") from exc
    update_cell = next(
        sheet.iter_rows(min_row=4, max_row=4, values_only=True)
    )[0]
    update_match = re.search(
        r"Updated on ([A-Za-z]+ \d{1,2}, \d{4})",
        str(update_cell or ""),
    )
    published_at = retrieved_at
    if update_match:
        try:
            published_at = datetime.strptime(
                update_match.group(1),
                "%B %d, %Y",
            ).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            published_at = retrieved_at
    header = list(next(sheet.iter_rows(min_row=5, max_row=5, values_only=True)))
    units = list(next(sheet.iter_rows(min_row=6, max_row=6, values_only=True)))
    requested = {
        value.casefold(): value
        for value in request["series"].split("|")
    }
    columns = {
        index: str(name).strip()
        for index, name in enumerate(header)
        if name is not None and str(name).strip().casefold() in requested
    }
    if len(columns) != len(requested):
        raise ValueError("public_market_series_not_found")
    values: dict[int, list[tuple[str, float]]] = {index: [] for index in columns}
    for row in sheet.iter_rows(min_row=7, values_only=True):
        period = str(row[0] or "").strip()
        if not re.fullmatch(r"\d{4}M\d{2}", period):
            continue
        for index in columns:
            value = row[index] if index < len(row) else None
            if isinstance(value, (int, float)):
                values[index].append((period, float(value)))
    observations: list[dict[str, Any]] = []
    for index, series_name in columns.items():
        points = values[index][-24:]
        unit = str(units[index] or "").strip() if index < len(units) else ""
        prior: float | None = None
        for period, value in points:
            direction = (
                "unknown"
                if prior is None
                else "increase" if value > prior
                else "decrease" if value < prior
                else "stable"
            )
            year, month = period.split("M", 1)
            effective = f"{year}-{month}-01T00:00:00+00:00"
            measurement = {
                "kind": "commodity_benchmark_price",
                "direction": direction,
                "series": series_name,
                "period": period,
                "value": value,
                "uom": unit or None,
                "currency": "USD" if "$" in unit else None,
            }
            observations.append(govern_external_observation(
                source_id="world_bank_pink_sheet",
                source_record_id=f"{series_name}:{period}",
                signal_type=request["signal_type"],
                subject_id=f"world-bank-commodity:{series_name.casefold()}",
                measurement=measurement,
                geography="global_benchmark",
                effective_from=effective,
                effective_to=None,
                published_at=published_at,
                available_at=published_at,
                retrieved_at=retrieved_at,
            ))
            prior = value
    return observations[:_MAX_ROWS]


def _latest_revision(
    db,
    *,
    tenant_id: str,
    source_id: str,
    request_key: str,
) -> dict[str, Any] | None:
    try:
        row = db.execute(text(
            f"SELECT revision_number, outcome, etag, last_modified, normalized_json, "
            f"expires_at FROM {_TABLE} WHERE tenant_id=:tenant AND source_id=:source "
            "AND request_key=:request_key ORDER BY revision_number DESC LIMIT 1"
        ), {
            "tenant": tenant_id,
            "source": source_id,
            "request_key": request_key,
        }).mappings().first()
    except Exception:
        return None
    return dict(row) if row else None


def _parse_time(value: Any) -> datetime | None:
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return _utc(parsed)
    except (TypeError, ValueError):
        return None


def _append_revision(
    db,
    *,
    tenant_id: str,
    source_id: str,
    request_key: str,
    revision_number: int,
    request: dict[str, str],
    outcome: str,
    now: datetime,
    expires_at: datetime,
    source_policy: dict[str, Any],
    http_status: int | None = None,
    etag: str | None = None,
    last_modified: str | None = None,
    content_sha256: str | None = None,
    normalized: list[dict[str, Any]] | None = None,
    error_code: str | None = None,
) -> None:
    db.execute(text(
        f"INSERT INTO {_TABLE} "
        "(id, tenant_id, source_id, request_key, revision_number, request_json, "
        "outcome, http_status, etag, last_modified, content_sha256, "
        "normalized_json, source_policy_json, error_code, retrieved_at, expires_at) "
        "VALUES (:id,:tenant,:source,:request_key,:revision,:request_json,:outcome,"
        ":http_status,:etag,:last_modified,:content_sha256,:normalized_json,"
        ":source_policy_json,:error_code,:retrieved_at,:expires_at)"
    ), {
        "id": uuid.uuid4().hex,
        "tenant": tenant_id,
        "source": source_id,
        "request_key": request_key,
        "revision": revision_number,
        "request_json": json.dumps(request, sort_keys=True),
        "outcome": outcome,
        "http_status": http_status,
        "etag": etag,
        "last_modified": last_modified,
        "content_sha256": content_sha256,
        "normalized_json": (
            json.dumps(normalized, sort_keys=True) if normalized is not None else None
        ),
        "source_policy_json": json.dumps(source_policy, sort_keys=True),
        "error_code": error_code,
        "retrieved_at": now,
        "expires_at": expires_at,
    })
    db.commit()


def group_contradictions(
    observations: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for observation in observations:
        key = (
            str(observation.get("subject_id") or ""),
            str(observation.get("signal_type") or ""),
        )
        groups.setdefault(key, []).append(observation)
    return [
        {
            "subject_id": subject,
            "signal_type": signal,
            **resolve_contradictions(rows),
        }
        for (subject, signal), rows in sorted(groups.items())
    ]


def fetch_public_market_source(
    db,
    *,
    tenant_id: str,
    source_id: str,
    query: dict[str, Any],
    now: datetime | None = None,
    transport: Transport | None = None,
    enabled: bool | None = None,
) -> dict[str, Any]:
    """Fetch, normalize and revision an approved source without raising provider errors."""
    tenant = str(tenant_id or "").strip()
    if not tenant:
        raise ValueError("public_market_tenant_required")
    allow_live = enabled if enabled is not None else (
        os.getenv("PUBLIC_MARKET_FETCH_ENABLED", "0").strip().lower()
        in {"1", "true", "yes", "on"}
    )
    if not allow_live:
        return {
            "source_id": source_id,
            "outcome": "disabled",
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    source = load_market_source_registry().get(str(source_id))
    if source is None:
        raise ValueError("external_market_source_not_registered")
    profile = source.get("fetch_profile")
    if not isinstance(profile, dict):
        return {
            "source_id": source_id,
            "outcome": "unsupported",
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    kind = str(profile["kind"])
    if kind == "cpsc_recalls_json":
        request = _normalize_cpsc_query(query)
        provider_params = request
    elif kind == "nws_active_alerts_json":
        request = _normalize_nws_query(query)
        provider_params = request
    elif kind == "usgs_mcs_sciencebase_csv":
        request = _normalize_usgs_query(query)
        provider_params = {}
    elif kind == "world_bank_pink_sheet_xlsx":
        request = _normalize_world_bank_query(query)
        provider_params = {}
    else:
        raise ValueError("external_market_source_fetch_kind_invalid")
    request_key = _request_key(request)
    stamp = _utc(now)
    latest = _latest_revision(
        db,
        tenant_id=tenant,
        source_id=source_id,
        request_key=request_key,
    )
    expiry = _parse_time(latest.get("expires_at")) if latest else None
    if latest and expiry and expiry > stamp and latest.get("normalized_json"):
        observations = json.loads(latest["normalized_json"])
        return {
            "source_id": source_id,
            "outcome": "cache_hit",
            "revision_number": int(latest["revision_number"]),
            "observations": observations,
            "contradictions": group_contradictions(observations),
            "fresh_until": expiry.isoformat(),
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    headers: dict[str, str] = {}
    if latest and latest.get("etag"):
        headers["If-None-Match"] = str(latest["etag"])
    if latest and latest.get("last_modified"):
        headers["If-Modified-Since"] = str(latest["last_modified"])
    ttl = max(60, int(profile.get("ttl_seconds") or 3600))
    fresh_until = stamp + timedelta(seconds=ttl)
    revision = int(latest.get("revision_number") or 0) + 1 if latest else 1
    source_policy = {
        "source_system": source["source_system"],
        "licence_id": source["licence_id"],
        "licence_url": source["licence_url"],
        "permitted_uses": source["permitted_uses"],
        "measurement_scope": source["measurement_scope"],
    }
    release: dict[str, Any] = {}
    try:
        selected_transport = transport or _default_transport
        if kind == "usgs_mcs_sciencebase_csv":
            response, release = _usgs_data_response(
                profile,
                transport=selected_transport,
                conditional_headers=headers,
            )
        else:
            response = selected_transport(
                str(profile["url"]),
                provider_params,
                headers,
                float(profile.get("timeout_seconds") or 8),
            )
    except ValueError:
        return {
            "source_id": source_id,
            "outcome": "malformed",
            "error_code": "provider_metadata_invalid",
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    except Exception:
        return {
            "source_id": source_id,
            "outcome": "unavailable",
            "error_code": "provider_unavailable",
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    if len(response.body) > int(profile.get("max_response_bytes") or 2_000_000):
        return {
            "source_id": source_id,
            "outcome": "malformed",
            "error_code": "response_too_large",
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    normalized: list[dict[str, Any]]
    outcome = "observed"
    if response.status_code == 304 and latest and latest.get("normalized_json"):
        normalized = json.loads(latest["normalized_json"])
        outcome = "not_modified"
    elif response.status_code == 429:
        return {
            "source_id": source_id,
            "outcome": "rate_limited",
            "retry_after": response.headers.get("retry-after"),
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    elif response.status_code < 200 or response.status_code >= 300:
        return {
            "source_id": source_id,
            "outcome": "unavailable",
            "http_status": response.status_code,
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    else:
        try:
            if kind == "cpsc_recalls_json":
                normalized = _cpsc_observations(
                    json.loads(response.body.decode("utf-8")),
                    retrieved_at=stamp.isoformat(),
                )
            elif kind == "nws_active_alerts_json":
                normalized = _nws_observations(
                    json.loads(response.body.decode("utf-8")),
                    request=request,
                    retrieved_at=stamp.isoformat(),
                )
            elif kind == "usgs_mcs_sciencebase_csv":
                normalized = _usgs_observations(
                    response.body,
                    request=request,
                    retrieved_at=stamp.isoformat(),
                    release=release,
                )
            else:
                normalized = _world_bank_observations(
                    response.body,
                    request=request,
                    retrieved_at=stamp.isoformat(),
                )
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError):
            return {
                "source_id": source_id,
                "outcome": "malformed",
                "error_code": "provider_payload_invalid",
                "authority": "advisory_only",
                "execution_allowed": False,
            }
    content_hash = hashlib.sha256(response.body).hexdigest() if response.body else None
    try:
        _append_revision(
            db,
            tenant_id=tenant,
            source_id=source_id,
            request_key=request_key,
            revision_number=revision,
            request=request,
            outcome=outcome,
            now=stamp,
            expires_at=fresh_until,
            source_policy=source_policy,
            http_status=response.status_code,
            etag=response.headers.get("etag") or (latest or {}).get("etag"),
            last_modified=(
                response.headers.get("last-modified")
                or (latest or {}).get("last_modified")
            ),
            content_sha256=content_hash,
            normalized=normalized,
        )
    except Exception:
        db.rollback()
        return {
            "source_id": source_id,
            "outcome": "unavailable",
            "error_code": "fetch_revision_store_unavailable",
            "authority": "advisory_only",
            "execution_allowed": False,
        }
    return {
        "source_id": source_id,
        "outcome": outcome,
        "revision_number": revision,
        "observations": normalized,
        "contradictions": group_contradictions(normalized),
        "fresh_until": fresh_until.isoformat(),
        "authority": "advisory_only",
        "can_establish_sku_exposure": False,
        "execution_allowed": False,
    }
